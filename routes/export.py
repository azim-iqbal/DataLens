import json
import os
from datetime import datetime
from html import escape

from fastapi import APIRouter, Body, HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from services.dataset_service import REPORT_DIR, ensure_data_dirs
from services.dictionary_cache import get_result, update_result


router = APIRouter()


def _safe_filename(name: str) -> str:
    return os.path.basename(name or "dataset").replace(" ", "_")


def _fairness_text(flag: dict | None) -> str:
    if not flag:
        return ""
    groq = flag.get("groq_verification") or {}
    verification = ""
    if groq:
        verification = f" Groq verification: {groq.get('verdict', 'Uncertain')} - {groq.get('reason', '')}"
    return f"{flag.get('eu_ai_act_article') or ''} {flag.get('reason') or ''}{verification}".strip()


def _p(text) -> str:
    return escape(str(text or ""))


def _draw_watermark(canvas, doc) -> None:
    canvas.saveState()
    width, height = letter
    canvas.setFillColor(colors.HexColor("#f7f3ea"))
    canvas.rect(0, 0, width, height, fill=1, stroke=0)
    canvas.setStrokeColor(colors.HexColor("#e6ded0"))
    canvas.setLineWidth(0.6)
    for x in range(36, int(width), 72):
        canvas.line(x, 0, x, height)
    for y in range(36, int(height), 72):
        canvas.line(0, y, width, y)

    canvas.translate(width / 2, height / 2)
    canvas.rotate(32)
    canvas.setFillColor(colors.Color(0.06, 0.06, 0.07, alpha=0.045))
    canvas.setFont("Helvetica-Bold", 76)
    canvas.drawCentredString(0, 0, "DataLens")
    canvas.setFillColor(colors.Color(0.0, 0.65, 0.78, alpha=0.08))
    canvas.setFont("Helvetica-Bold", 20)
    canvas.drawCentredString(0, -30, "DATA DICTIONARY")
    canvas.restoreState()

    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#111214"))
    canvas.roundRect(36, height - 42, 24, 24, 6, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#fffaf0"))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawCentredString(48, height - 27, "DL")
    canvas.setFillColor(colors.HexColor("#4c473f"))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(66, height - 28, "DataLens Studio")
    canvas.setFillColor(colors.HexColor("#777166"))
    canvas.drawRightString(width - 36, 22, f"Page {doc.page}")
    canvas.restoreState()


def create_dictionary_pdf(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.pdf")
    doc = SimpleDocTemplate(
        path,
        pagesize=letter,
        rightMargin=0.62 * inch,
        leftMargin=0.62 * inch,
        topMargin=0.78 * inch,
        bottomMargin=0.58 * inch,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="StudioTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=34,
        leading=36,
        textColor=colors.HexColor("#111214"),
        spaceAfter=14,
    ))
    styles.add(ParagraphStyle(
        name="StudioH2",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=20,
        textColor=colors.HexColor("#111214"),
        spaceAfter=8,
    ))
    styles.add(ParagraphStyle(
        name="StudioBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.6,
        leading=13,
        textColor=colors.HexColor("#332f29"),
    ))
    styles.add(ParagraphStyle(
        name="StudioCode",
        parent=styles["Code"],
        fontName="Courier",
        fontSize=7.6,
        leading=10,
        textColor=colors.HexColor("#fffaf0"),
        backColor=colors.HexColor("#111214"),
        borderPadding=7,
    ))
    story = []

    readiness = result.get("readiness", {})
    metadata = result.get("metadata", {})
    story.append(Paragraph("DataLens Data Dictionary", styles["StudioTitle"]))
    story.append(Paragraph(
        f"Dataset <b>{_p(metadata.get('filename', 'Unknown'))}</b> was analysed on "
        f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}.",
        styles["StudioBody"],
    ))
    story.append(Spacer(1, 18))
    cover_rows = [
        ["Rows", "Columns", "Readiness", "Grade"],
        [
            str(metadata.get("row_count", 0)),
            str(metadata.get("column_count", 0)),
            str(readiness.get("score", 0)),
            str(readiness.get("grade", "N/A")),
        ],
    ]
    cover = Table(cover_rows, colWidths=[1.15 * inch] * 4)
    cover.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111214")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#fffaf0")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#fffaf0")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#111214")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, 1), 22),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#111214")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d8d0c2")),
    ]))
    story.append(cover)
    story.append(Spacer(1, 16))
    story.append(Paragraph(_p(readiness.get("summary", "")), styles["StudioBody"]))
    story.append(PageBreak())

    for profile in result.get("profiles", []):
        story.append(Paragraph(_p(profile.get("display_name") or profile.get("column_name")), styles["StudioH2"]))
        card_rows = [
            ["Column", _p(profile.get("column_name")), "Type", _p(profile.get("dtype"))],
            ["Nulls", f"{_p(profile.get('null_pct'))}%", "Unique", _p(profile.get("unique_count"))],
            ["Confidence", _p(profile.get("confidence")), "ML Concern", _p(profile.get("data_quality_flag"))],
            ["Description", Paragraph(_p(profile.get("description") or ""), styles["StudioBody"]), "", ""],
        ]
        if profile.get("anomaly_note"):
            card_rows.append(["Anomaly", Paragraph(_p(profile.get("anomaly_note")), styles["StudioBody"]), "", ""])
        if profile.get("fairness_flag"):
            card_rows.append(["EU AI Act", Paragraph(_p(_fairness_text(profile.get("fairness_flag"))), styles["StudioBody"]), "", ""])
        card = Table(card_rows, colWidths=[0.9 * inch, 2.25 * inch, 0.9 * inch, 1.35 * inch])
        card_styles = [
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fffaf0")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6b665e")),
            ("TEXTCOLOR", (2, 0), (2, 2), colors.HexColor("#6b665e")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, 2), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("SPAN", (1, 3), (3, 3)),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#d8d0c2")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e7dfd2")),
        ]
        for row_index in range(4, len(card_rows)):
            card_styles.append(("SPAN", (1, row_index), (3, row_index)))
        card.setStyle(TableStyle(card_styles))
        story.append(card)
        story.append(Spacer(1, 14))

    story.append(PageBreak())
    story.append(Paragraph("Relationships", styles["StudioH2"]))
    relationship_rows = [["Column A", "Column B", "Correlation", "Type", "Note"]]
    for rel in result.get("relationships", []):
        relationship_rows.append([_p(rel.get("col_a")), _p(rel.get("col_b")), _p(rel.get("correlation")), _p(rel.get("type")), Paragraph(_p(rel.get("note")), styles["StudioBody"])])
    for rel in result.get("redundant_columns", []):
        relationship_rows.append([_p(rel.get("col_a")), _p(rel.get("col_b")), f"{_p(rel.get('match_pct'))}% match", "Near duplicate", Paragraph(_p(rel.get("note")), styles["StudioBody"])])
    relationship_table = Table(relationship_rows, repeatRows=1, colWidths=[1.05 * inch, 1.05 * inch, 0.8 * inch, 1.05 * inch, 1.75 * inch])
    relationship_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111214")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#fffaf0")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#fffaf0")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d8d0c2")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(relationship_table)
    story.append(Spacer(1, 18))

    story.append(Paragraph("Query Suggestions", styles["StudioH2"]))
    for query in result.get("query_suggestions", []):
        story.append(Paragraph(_p(query.get("question", "")), styles["StudioBody"]))
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"Pandas: {_p(query.get('pandas_query', ''))}", styles["StudioCode"]))
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"SQL: {_p(query.get('sql_query', ''))}", styles["StudioCode"]))
        story.append(Spacer(1, 12))

    doc.build(story, onFirstPage=_draw_watermark, onLaterPages=_draw_watermark)
    return path


def create_dictionary_excel(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"
    ws.freeze_panes = "A2"
    headers = [
        "Column Name",
        "Display Name",
        "Type",
        "Null %",
        "Unique Count",
        "Description",
        "Confidence",
        "Anomaly Note",
        "Fairness Flag",
        "Groq Verification",
        "ML Concern",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="111214")
    header_font = Font(bold=True, color="FFFAF0")
    thin_side = Side(style="thin", color="D8D0C2")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28

    red = PatternFill("solid", fgColor="FAD7D7")
    amber = PatternFill("solid", fgColor="FBE6B5")
    green = PatternFill("solid", fgColor="E7F3C5")
    default_fill = PatternFill("solid", fgColor="FFFAF0")

    for profile in result.get("profiles", []):
        ws.append([
            profile.get("column_name"),
            profile.get("display_name"),
            profile.get("dtype"),
            profile.get("null_pct"),
            profile.get("unique_count"),
            profile.get("description"),
            profile.get("confidence"),
            profile.get("anomaly_note"),
            _fairness_text(profile.get("fairness_flag")),
            (profile.get("fairness_flag") or {}).get("verification_status"),
            profile.get("data_quality_flag"),
        ])
        row = ws.max_row
        if profile.get("fairness_flag"):
            fill = red
        elif profile.get("anomaly_note"):
            fill = amber
        elif profile.get("confidence") == "Confirmed":
            fill = green
        else:
            fill = default_fill

        for cell in ws[row]:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in {1, 2, 3, 4, 5, 7, 10, 11}:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 54

        if profile.get("fairness_flag"):
            for cell in ws[row]:
                cell.fill = red

    for column_cells in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 4, 64)
        ws.column_dimensions[column_cells[0].column_letter].width = width
    ws.column_dimensions["F"].width = 54
    ws.column_dimensions["H"].width = 44
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["J"].width = 24

    wb.save(path)
    return path


def create_dictionary_json(session_id: str, result: dict) -> str:
    ensure_data_dirs()
    path = os.path.join(REPORT_DIR, f"{session_id}_dictionary.json")
    with open(path, "w", encoding="utf-8") as json_file:
        json.dump(result.get("profiles", []), json_file, indent=2, default=str)
    return path


def export_result(session_id: str, export_format: str, result: dict) -> FileResponse:
    export_format = export_format.lower()
    if export_format == "pdf":
        path = create_dictionary_pdf(session_id, result)
        return FileResponse(path, media_type="application/pdf", filename=f"datalens_{session_id}.pdf")
    if export_format == "excel":
        path = create_dictionary_excel(session_id, result)
        return FileResponse(path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"datalens_{session_id}.xlsx")
    if export_format == "json":
        path = create_dictionary_json(session_id, result)
        return FileResponse(path, media_type="application/json", filename=f"datalens_{session_id}.json")
    raise HTTPException(status_code=400, detail="Format must be pdf, excel, or json")


@router.get("/export/{session_id}/{format}")
async def export_cached_result(session_id: str, format: str):
    result = get_result(session_id)
    if not result:
        raise HTTPException(status_code=404, detail="Session result not found")
    return export_result(session_id, format, result)


@router.post("/export/{session_id}/{format}")
async def export_edited_result(session_id: str, format: str, result: dict = Body(...)):
    update_result(session_id, result)
    return export_result(session_id, format, result)
